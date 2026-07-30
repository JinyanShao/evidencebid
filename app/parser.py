from __future__ import annotations

from collections.abc import Iterable
from pathlib import Path
from uuid import uuid4

from docling.document_converter import DocumentConverter
from openpyxl import load_workbook
from pypdf import PdfReader

from app.models import ExtractedBlock, SourceLocation


SUPPORTED_EXTENSIONS = {".pdf", ".docx", ".xlsx"}


class DocumentParseError(ValueError):
    """An error that can be safely shown to an EvidenceBid user."""


def parse_document(path: Path) -> list[ExtractedBlock]:
    """Convert a supported text document with Docling and retain source provenance."""
    suffix = path.suffix.lower()
    if suffix not in SUPPORTED_EXTENSIONS:
        raise DocumentParseError("Only text PDFs, DOCX, and XLSX files are supported in this version.")
    if path.stat().st_size == 0:
        raise DocumentParseError("This file is empty. Upload a file that contains text or cell values.")

    try:
        if suffix == ".pdf":
            return _pdf_blocks(path)
        result = DocumentConverter().convert(path)
        document = result.document
    except Exception as exc:  # Docling exposes parser-specific exceptions inconsistently across formats.
        message = str(exc).lower()
        if suffix == ".pdf" and any(word in message for word in ("password", "encrypted", "decrypt")):
            raise DocumentParseError("This PDF is password-protected. Upload an unlocked text-based PDF.") from exc
        raise DocumentParseError(
            "We could not read this file. This version supports text-based PDFs, DOCX, and XLSX files; "
            "scanned, damaged, and password-protected files are not supported."
        ) from exc

    blocks = _xlsx_blocks(path) if suffix == ".xlsx" else list(_blocks_from_docling(document, suffix))
    if not blocks:
        if suffix == ".pdf":
            raise DocumentParseError(
                "No selectable text was found. This appears to be a scanned PDF; OCR is not supported in this version."
            )
        raise DocumentParseError("No readable text or cell values were found in this file.")
    return blocks


def _pdf_blocks(path: Path) -> list[ExtractedBlock]:
    """Extract text PDFs without OCR or model downloads, retaining exact page numbers."""
    try:
        reader = PdfReader(path)
    except Exception as exc:
        raise DocumentParseError("This PDF could not be opened. Upload an unlocked, text-based PDF.") from exc

    blocks: list[ExtractedBlock] = []
    for page_number, page in enumerate(reader.pages, start=1):
        text = (page.extract_text() or "").strip()
        if text:
            blocks.append(
                ExtractedBlock(
                    id=str(uuid4()),
                    kind="page text",
                    text=text,
                    source=SourceLocation(label=f"Page {page_number}", page=page_number),
                )
            )
    return blocks


def _blocks_from_docling(document: object, suffix: str) -> Iterable[ExtractedBlock]:
    """Use Docling's document items, avoiding private parser internals.

    The document model exposes labelled text/table items and provenance.  The defensive
    accessors below keep our app tolerant of minor Docling model changes.
    """
    iterator = getattr(document, "iterate_items", None)
    if not callable(iterator):
        markdown = document.export_to_markdown()  # type: ignore[attr-defined]
        yield ExtractedBlock(id=str(uuid4()), kind="text", text=markdown, source=SourceLocation(label="Document"))
        return

    for item, _level in iterator():
        text = _item_text(item, document)
        if not text:
            continue
        kind = str(getattr(item, "label", "text")).lower()
        provenance = _source_location(item, suffix)
        yield ExtractedBlock(id=str(uuid4()), kind=kind, text=text, source=provenance)


def _item_text(item: object, document: object) -> str:
    if hasattr(item, "export_to_markdown"):
        text = item.export_to_markdown(doc=document)  # type: ignore[attr-defined]
    else:
        text = getattr(item, "text", "")
    return str(text).strip()


def _source_location(item: object, suffix: str) -> SourceLocation:
    prov = (getattr(item, "prov", None) or [None])[0]
    page_no = getattr(prov, "page_no", None) if prov else None
    label = f"Page {page_no}" if page_no else "Document"
    section = getattr(item, "text", None) if str(getattr(item, "label", "")).lower() in {"section_header", "title"} else None
    if suffix == ".xlsx":
        # Docling preserves spreadsheet content; a worksheet name is included in the
        # document text where available. The app falls back to an explicit worksheet label.
        sheet_name = getattr(prov, "sheet_name", None) if prov else None
        return SourceLocation(label=f"Worksheet {sheet_name or 'content'}", sheet=sheet_name, section=section)
    return SourceLocation(label=label, page=page_no, section=section)


def _xlsx_blocks(path: Path) -> list[ExtractedBlock]:
    """Preserve the worksheet and cell range that Docling's generic provenance omits."""
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise DocumentParseError("This workbook could not be opened. Upload a valid, unprotected XLSX file.") from exc

    blocks: list[ExtractedBlock] = []
    for worksheet in workbook.worksheets:
        for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            values = [str(value).strip() for value in row if value is not None and str(value).strip()]
            if not values:
                continue
            cell_range = f"A{row_number}:{_column_name(len(row))}{row_number}"
            blocks.append(
                ExtractedBlock(
                    id=str(uuid4()),
                    kind="worksheet row",
                    text=" | ".join(values),
                    source=SourceLocation(label=f"Worksheet {worksheet.title} · {cell_range}", sheet=worksheet.title),
                )
            )
    return blocks


def _column_name(index: int) -> str:
    letters = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters
